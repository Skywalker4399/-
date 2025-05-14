import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 定义 Excel 文件路径
file_path = './output/花垣沉降预测数据_汇总_0512.xlsx'

# 读取 Excel 文件
data = pd.ExcelFile(file_path).parse(['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4'])

# 加载工作簿
wb = load_workbook(file_path)

# 获取工作表
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']
sheet3 = wb['Sheet3']
sheet4 = wb['Sheet4']

#X操作
sheet2.cell(row=1, column=1).value = '日期'
# 获取 Sheet1 中 JC01 范围的日期列的数据并转置复制到 Sheet2 第一行，从第二列开始复制
jc01_dates = data['Sheet1'][data['Sheet1']['点名'] == 'JC01']['日期'].tolist()
for col, value in enumerate(jc01_dates, start=2):
    sheet2.cell(row=1, column=col).value = value

# 写入点名 JC01 - JC12 到 Sheet2 的 A2 - A13 单元格
for row, point_name in enumerate([f'JC{i:02}' for i in range(1, 13)], start=2):
    sheet2.cell(row=row, column=1).value = point_name

# 提取 Sheet1 中 JC01 - JC12 对应的 X_累计位移(mm)数据
selected_data = data['Sheet1'][data['Sheet1']['点名'].isin([f'JC{i:02}' for i in range(1, 13)])][['点名', '日期', 'X_累计位移(mm)']]

# 使用 pivot_table 方法创建透视表
pivot_table = selected_data.pivot_table(index='点名', columns='日期', values='X_累计位移(mm)', aggfunc='first').reset_index()

# 将数据按行写入 Sheet2 从第 2 行 B 列开始
for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=False), 2):
    for c_idx, value in enumerate(row[1:], start=2):
        sheet2.cell(row=r_idx, column=c_idx).value = value

#Y操作
sheet3.cell(row=1, column=1).value = '日期'
# 获取 Sheet1 中 JC01 范围的日期列的数据并转置复制到 Sheet2 第一行，从第二列开始复制
jc01_dates = data['Sheet1'][data['Sheet1']['点名'] == 'JC01']['日期'].tolist()
for col, value in enumerate(jc01_dates, start=2):
    sheet3.cell(row=1, column=col).value = value

# 写入点名 JC01 - JC12 到 Sheet2 的 A2 - A13 单元格
for row, point_name in enumerate([f'JC{i:02}' for i in range(1, 13)], start=2):
    sheet3.cell(row=row, column=1).value = point_name

# 提取 Sheet1 中 JC01 - JC12 对应的 Y_累计位移(mm)数据
selected_data = data['Sheet1'][data['Sheet1']['点名'].isin([f'JC{i:02}' for i in range(1, 13)])][['点名', '日期', 'Y_累计位移(mm)']]

# 使用 pivot_table 方法创建透视表
pivot_table = selected_data.pivot_table(index='点名', columns='日期', values='Y_累计位移(mm)', aggfunc='first').reset_index()

# 将数据按行写入 Sheet2 从第 2 行 B 列开始
for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=False), 2):
    for c_idx, value in enumerate(row[1:], start=2):
        sheet3.cell(row=r_idx, column=c_idx).value = value

#Z操作
sheet4.cell(row=1, column=1).value = '日期'
# 获取 Sheet1 中 JC01 范围的日期列的数据并转置复制到 Sheet2 第一行，从第二列开始复制
jc01_dates = data['Sheet1'][data['Sheet1']['点名'] == 'JC01']['日期'].tolist()
for col, value in enumerate(jc01_dates, start=2):
    sheet4.cell(row=1, column=col).value = value

# 写入点名 JC01 - JC12 到 Sheet2 的 A2 - A13 单元格
for row, point_name in enumerate([f'JC{i:02}' for i in range(1, 13)], start=2):
    sheet4.cell(row=row, column=1).value = point_name

# 提取 Sheet1 中 JC01 - JC12 对应的 Z_累计下沉(mm)数据
selected_data = data['Sheet1'][data['Sheet1']['点名'].isin([f'JC{i:02}' for i in range(1, 13)])][['点名', '日期', 'Z_累计下沉(mm)']]

# 使用 pivot_table 方法创建透视表
pivot_table = selected_data.pivot_table(index='点名', columns='日期', values='Z_累计下沉(mm)', aggfunc='first').reset_index()

# 将数据按行写入 Sheet2 从第 2 行 B 列开始
for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=False), 2):
    for c_idx, value in enumerate(row[1:], start=2):
        sheet4.cell(row=r_idx, column=c_idx).value = value

# 保存修改后的 Excel 文件
wb.save('./output/花垣沉降_成图数据_0512.xlsx')
    
